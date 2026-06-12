#!/usr/bin/env python3
"""Create and maintain a receipt-expense spreadsheet.

Subcommands: init, append, sort, total.
Rows are passed as JSON so quoting/locale characters survive cleanly.
Each data row is [issue_date, description, total_value, filename].

Requires openpyxl (pip install openpyxl --break-system-packages).
"""
import argparse
import json
import sys
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


FONT = "Arial"
VALUE_COL = 3  # 1-based: the "Total value" column


def _style_data_row(ws, row_idx, currency_format):
    for c in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=c).font = Font(name=FONT)
    ws.cell(row=row_idx, column=VALUE_COL).number_format = currency_format


def cmd_init(args):
    headers = json.loads(args.headers)
    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet
    ws.append(headers)
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="1F4E78")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [16, 48, 14, 32]
    for i in range(len(headers)):
        ws.column_dimensions[chr(65 + i)].width = widths[i] if i < len(widths) else 18
    ws.freeze_panes = "A2"
    wb.save(args.path)
    print(json.dumps({"status": "ok", "action": "init", "path": args.path,
                      "headers": headers}))


def cmd_append(args):
    rows = json.loads(args.rows)
    wb = load_workbook(args.path)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active
    for r in rows:
        ws.append(r)
        _style_data_row(ws, ws.max_row, args.currency_format)
    wb.save(args.path)
    print(json.dumps({"status": "ok", "action": "append",
                      "added": len(rows), "total_rows": ws.max_row - 1}))


def cmd_sort(args):
    wb = load_workbook(args.path)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active
    data = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in r)]

    def keyfn(row):
        try:
            return (0, datetime.strptime(str(row[0]), args.date_format))
        except Exception:
            return (1, datetime.max)  # unparseable dates sink to the bottom

    data.sort(key=keyfn)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for r in data:
        ws.append(r)
        _style_data_row(ws, ws.max_row, args.currency_format)
    wb.save(args.path)
    print(json.dumps({"status": "ok", "action": "sort", "rows": len(data)}))


def cmd_total(args):
    wb = load_workbook(args.path)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active
    last = ws.max_row
    if last < 2:
        print(json.dumps({"status": "noop", "reason": "no data rows"}))
        return
    col = chr(64 + VALUE_COL)
    ws.append([args.label, None, f"=SUM({col}2:{col}{last})", None])
    rn = ws.max_row
    for c in range(1, ws.max_column + 1):
        ws.cell(row=rn, column=c).font = Font(name=FONT, bold=True)
    ws.cell(row=rn, column=VALUE_COL).number_format = args.currency_format
    wb.save(args.path)
    print(json.dumps({"status": "ok", "action": "total", "sum_row": rn,
                      "note": "recalc with LibreOffice if a cached value is needed"}))


def main():
    p = argparse.ArgumentParser(description="Receipt expense sheet helper")
    sub = p.add_subparsers(dest="cmd", required=True)

    common_currency = dict(default="R$ #,##0.00",
                           help="number format for the value column")
    common_sheet = dict(default="Despesas", help="worksheet name")

    pi = sub.add_parser("init")
    pi.add_argument("--path", required=True)
    pi.add_argument("--headers",
                    default='["Data de Emissão","Descrição","Valor Total","Arquivo"]')
    pi.add_argument("--sheet", **common_sheet)
    pi.set_defaults(func=cmd_init)

    pa = sub.add_parser("append")
    pa.add_argument("--path", required=True)
    pa.add_argument("--rows", required=True, help="JSON list of [date,desc,total,file]")
    pa.add_argument("--currency-format", **common_currency)
    pa.add_argument("--sheet", **common_sheet)
    pa.set_defaults(func=cmd_append)

    ps = sub.add_parser("sort")
    ps.add_argument("--path", required=True)
    ps.add_argument("--date-format", default="%d/%m/%Y")
    ps.add_argument("--currency-format", **common_currency)
    ps.add_argument("--sheet", **common_sheet)
    ps.set_defaults(func=cmd_sort)

    pt = sub.add_parser("total")
    pt.add_argument("--path", required=True)
    pt.add_argument("--label", default="TOTAL")
    pt.add_argument("--currency-format", **common_currency)
    pt.add_argument("--sheet", **common_sheet)
    pt.set_defaults(func=cmd_total)

    args = p.parse_args()
    try:
        args.func(args)
    except Exception as e:
        print(json.dumps({"status": "error", "error": str(e)}), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
